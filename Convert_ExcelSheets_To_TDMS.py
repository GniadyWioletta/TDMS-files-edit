from nptdms import TdmsWriter, RootObject, GroupObject, ChannelObject 
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import os


if __name__ == "__main__":
    
    szybciej = False
    path = ""
    path_excel = path + "\\" + "AllData.xlsx"
    new_path = os.path.join(path, "tdms")
    
    already_done = []
    
    try:
        os.mkdir(new_path) 
    except:
        new_path = new_path
        
    excel_file = openpyxl.open(path_excel)
        
    for sht_i in range(len(excel_file.sheetnames)):
        
        found = False
        sheet = excel_file.get_sheet_by_name(excel_file.sheetnames[sht_i])
        Name = sheet.cell(1,1).value
        
        for file_done in already_done:
            if file_done == Name:
                found = True
                
        if found:
            continue
        
        print(Name)
        
        for i in range(1, 50):
            if sheet.cell(row=i, column=1).value == "Time":
                data_row = i+2
                break
        
        arr_data = []
        arr_dataNames = []
        arr_unit = []
            
        for i in range(1, 50):
            if not sheet.cell(row=data_row-2, column=i).value is None:
                dataName = sheet.cell(data_row-2, i).value
                print(dataName)
                arr_dataNames.append(dataName)
                arr_unit.append(sheet.cell(data_row-1, i).value)
                
                if dataName == "Time":
                    value1 = float(sheet.cell(data_row, i).value)
                    value2 = float(sheet.cell(data_row+1, i).value)
                    sample_rate = value2 - value1
                    samples_1s = 1/sample_rate
            
                rng = sheet[sheet.cell(data_row, i).coordinate + ":" + sheet.cell(sheet.max_row, i).coordinate]
                
                #data = np.array([(i[0]) for i in rng])
                
                data = np.array([])
                for el in rng:
                    if isinstance(el[0].value, datetime):
                        val = float(str(el[0].value.day) + "." + str(el[0].value.year))
                        data = np.append(data, val)
                    else:
                        data = np.append(data, float(el[0].value))
                
                arr_data.append(data)
                
        root_object = RootObject()
        group_object = GroupObject("Measurement")
            
        print("Exporting")
        with TdmsWriter(new_path + "\\" + Name + ".tdms") as tdms_writer:
            for i in range(len(arr_data)):
                
                channel_object = ChannelObject("Measurement", arr_dataNames[i], arr_data[i], 
                                               properties={"NI_Unit": arr_unit[i],
                                                           "NI_ChannelName": arr_dataNames[i],
                                                           "wf_increment": sample_rate, 
                                                           "wf_samples": samples_1s,
                                                           "wf_start_offset": 0, 
                                                           "wf_time_pref": "relative",
                                                           "wf_xname": "Time",
                                                           "wf_xunit_string": "s",
                                                           "unit_string": arr_unit[i],})
            
                try:
                    if i == 0 :
                        tdms_writer.write_segment([root_object,group_object,channel_object])
                    else:
                        tdms_writer.write_segment([channel_object])
                except:
                    print("Błąd przy: " + str(Name) + " kanał: " + arr_dataNames[i])
                    print("Excel część wartości okreslił jako datę - np. 1.5625 to będzie sty. 25. Zmień wartość <szybciej> na False. Wtedy pójdzie po każdej komórce osobno i zmieni te wartości na liczby.")
         
        #already_done.append(Name)       
        
    print("Done")
            


            
