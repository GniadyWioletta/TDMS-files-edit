from nptdms import TdmsFile, TdmsWriter, RootObject, GroupObject, ChannelObject 
import openpyxl
import numpy as np
import pandas as pd
from xlrd import open_workbook


if __name__ == "__main__":
    
    Name = "file_name"
    path_excel = ""
    path = path_excel + "\\" + Name + ".xls"
                  
    #excel_file = openpyxl.open(path_excel)
    #df = pd.read_excel(io=path)
    #sheet = df.parse("Name")
    
    excel_file =  open_workbook(path_excel)
    sheet = excel_file.get_sheet_by_name(Name)
            
    force_data = np.array([])
    
    CH_time = 1
    CH_disp = 2
    CH_acc = 5
    row_start = 6

    value1 = float(sheet.cell(row_start, CH_time).value)
    value2 = float(sheet.cell(row_start+1, CH_time).value)
    sample_rate = value2 - value1
    samples_1s_excel = 1/sample_rate
            
    rng = sheet[sheet.cell(row_start, CH_disp).coordinate + ":" + sheet.cell(sheet.max_row, CH_disp).coordinate]
    disp_data = np.array([(i[0]) for i in rng])
    rng = sheet[sheet.cell(row_start, CH_acc).coordinate + ":" + sheet.cell(sheet.max_row, CH_acc).coordinate]
    acc_data = np.array([(i[0]) for i in rng])
            
    root_object = RootObject()
    group_object = GroupObject("Measurement")
    channel_object_acc = ChannelObject("Measurement", "Acc_Z", acc_data, properties={"Unit":"g","NI_ChannelName": "Acc_Z",})
    channel_object_disp = ChannelObject("Measurement", "Disp", disp_data, properties={"Unit":"V","NI_ChannelName": "Disp",})
    
    with TdmsWriter(path_excel + "\\" + Name) as tdms_writer:
        channel_object = ChannelObject("Measurement", "Acc_Z", acc_data, properties={
            "NI_Unit": "g",
            "NI_ChannelName": "ACC_Z",
            "wf_increment": float(sample_rate),
            "wf_samples": samples_1s_excel, # 12800
            "wf_start_offset": 0,
            "wf_time_pref": "relative",
            "wf_xname": "Time",
            "wf_xunit_string": "s",
            "unit_string": "g",
            })
        tdms_writer.write_segment([
            root_object,
            group_object,
            channel_object])
        
        channel_object = ChannelObject("Measurement", "disp", disp_data, properties={
            "NI_Unit": "mm",
            "NI_ChannelName": "disp",
            "wf_increment": float(sample_rate),
            "wf_samples": samples_1s_excel, # 12800
            "wf_start_offset": 0,
            "wf_time_pref": "relative",
            "wf_xname": "Time",
            "wf_xunit_string": "s",
            "unit_string": "mm",
            })
        tdms_writer.write_segment([
            root_object,
            group_object,
            channel_object])

            
            
            


            
