from nptdms import TdmsFile, TdmsWriter, RootObject, GroupObject, ChannelObject 
import openpyxl
import numpy as np
from datetime import datetime


if __name__ == "__main__":
    
    szybciej = False
        #sample_no = 2
    for sample_no in range(1,3): #Ostatnia wartość sie nie wykonuje wiec 1 i 2 tylko, jesli range(1,3)
        for sample_no_part in range(1,6):
    
            Name = "Sample_No_" + str(sample_no) + "_" + str(sample_no_part)
            path_tdms = "" + Name + ".tdms"
            path_excel = "" + Name + ".xlsx"
    
            tdms_file = TdmsFile.read(path_tdms)
            all_groups = tdms_file.groups()
            try:
                group = tdms_file["Measurement"]
            except:
                group = tdms_file[all_groups[1].name]
                all_group_channels = group.channels()

            all_group_channels = group.channels()
            i = -1
            for CH in all_group_channels:
                i = i+1
                if "_Z" in CH.name:
                    acc_CH = all_group_channels[i]
                elif "Mic" in CH.name:
                    Mic_CH = all_group_channels[i]
                elif "disp" in CH.name:
                    disp_CH = all_group_channels[i]
                if "Temp" in CH.name:
                    temp_CH = all_group_channels[i]
                    
            sample_rate = acc_CH.properties["wf_increment"]
            SampleNo_wf = acc_CH.properties["wf_samples"]
            samples_1s_tdms = 1/sample_rate
            
            acc_data = acc_CH[:]
            mic_data = acc_CH[:]
            disp_data = acc_CH[:]
            temp_data = acc_CH[:]
            
            excel_file = openpyxl.open(path_excel)
            sheet = excel_file.get_sheet_by_name(Name)
            
            force_row = 0
            force_col = 0
            
            force_data = np.array([])
            #Find Force
            for row in range(1, 5):
                for col in range(1, 10):
                    if not sheet.cell(row=row, column=col).value is None:
                        if "Force" in sheet.cell(row=row, column=col).value:
                            force_row = row
                            force_col = col
                            break
                if force_row !=0:
                    break
                
            value1 = float(sheet.cell(force_row+2, 1).value)
            value2 = float(sheet.cell(force_row+3, 1).value)
            sample_rate = value2 - value1
            samples_1s_excel = 1/sample_rate
            
            rng = sheet[sheet.cell(force_row+2, force_col).coordinate + ":" + sheet.cell(sheet.max_row, force_col).coordinate]
            
            if szybciej:
                force_data = np.array([(i[0]) for i in rng])
            else:
                for i in rng:
                    if isinstance(i[0].value, datetime):
                        val = float(str(i[0].value.day) + "." + str(i[0].value.year))
                        force_data = np.append(force_data, val)
                    else:
                        force_data = np.append(force_data, i[0].value)
            
            
            datacount_tdms = len(acc_data)
            datacount_excel = len(force_data)
            samples_diff = int(samples_1s_tdms/samples_1s_excel)-1
            #total_time = datacount_tdms/samples_1s_tdms
            #total_time = datacount_excel/samples_1s_excel
            
            acc_data = acc_data[0::samples_diff]
            mic_data = mic_data[0::samples_diff]
            disp_data = disp_data[0::samples_diff]
            temp_data = temp_data[0::samples_diff]
            
            acc_data = acc_data[0:datacount_excel]
            mic_data = mic_data[0:datacount_excel]
            disp_data = disp_data[0:datacount_excel]
            temp_data = temp_data[0:datacount_excel]

            root_object = RootObject()
            group_object = GroupObject("Measurement")
            channel_object_acc = ChannelObject("Measurement", "Acc_Z", acc_data, properties={"Unit":"g","NI_ChannelName": "Acc_Z",})
            channel_object_mic = ChannelObject("Measurement", "Mic", mic_data, properties={"Unit":"Pa","NI_ChannelName": "Mic",})
            channel_object_disp = ChannelObject("Measurement", "Disp", disp_data, properties={"Unit":"V","NI_ChannelName": "Disp",})
            channel_object_force = ChannelObject("Measurement", "Force", force_data, properties={"Unit":"N","NI_ChannelName": "Force",})
            channel_object_temp = ChannelObject("Measurement", "Temp", temp_data, properties={"Unit":"V","NI_ChannelName": "Temp",})

            try:
                with TdmsWriter(Name + "_changed.tdms") as tdms_writer:
                    tdms_writer.write_segment([root_object,group_object,channel_object_disp])
                    tdms_writer.write_segment([channel_object_acc])
                    tdms_writer.write_segment([channel_object_force])
                    tdms_writer.write_segment([channel_object_mic])
                    tdms_writer.write_segment([channel_object_temp])
            except:
                print("Excel część wartości okreslił jako datę - np. 1.5625 to będzie sty. 25. Zmień wartość <szybciej> na False. Wtedy pójdzie po każdej komórce osobno i zmieni te wartości na liczby.")
            
            
            
            
            


            
